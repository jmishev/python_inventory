import logging
import os
import sys
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename

from openpyxl import Workbook, utils, load_workbook
from openpyxl.styles import PatternFill, Font

import MySQL as ms

labels = {}
entries = {}
buttons = {}
good_list = {}
# dictionaries of good objects

root = 'it becomes a root object'
v = 'it becomes a tkinter variable'

directory = str(os.path.dirname(os.path.abspath(__file__)))
# logger
format_log = "%(levelname)s %(asctime)s - %(message)s"
logging.basicConfig(filename=directory + "/logging.txt",
                    level=logging.DEBUG, format=format_log, filemode='w')
logger = logging.getLogger()

class Goods:
    def __init__(self, name, quantity, price):
        self.name = name
        self.quantity = quantity
        self.price = price
        self.total_price = self.quantity * self.price


def close_with_x():
    """ Closing by closing main window by clicking on 'X' """
    if not ((entries[(1, 1)].get() and entries[(2, 1)].get()) or good_list):
        # If no data is on screen and no goods are created exits
        logger.info("If no data is on screen and no goods are created exits")
        sys.exit()
    else:
        if check_on_screen() == "Good to be entered":
            enter_good()
            logger.info("Exiting data exported to excel(close_with_x)")
            write_or_add()
            return "Good entered"
        elif good_list:
            logger.info("No onscreen entry but goods in the list (close_with_x)")
            write_or_add()
            return "Goods in the list entered no onscreen entered (close_with_x)"
    sys.exit()

def write_or_add():
    """Selecting the output file new or existing"""
    logger.info("Selecting the output file new or existing (write_or_add)")
    
    if  root.title() == "Програма за ревизия":
        # destination file is not preselected
        logger.info("Writing the data to excel before closing with button (write_or_add)")
        write_to_excel()
            # create/select file and write the data (exiting data will be over written)
        return "written to excel"

    else:
        add_to_excel()
        # destination file is not preselected
        logger.info("Adding the data to excel before closing with button(write_or_add)")
        return "added to excel"


def close_with_button():  # event is given by pushing the button 'Запази и излез'
    """Closing with button 'Запази и излез'"""
    logger.info("starting close with button (close_with_button)")
    x = check_on_screen()
    if x == "No data" and good_list != {}:  # No data onscreen and no goods entered yet
        write_or_add()
        return "Closing with button no onscreen"
        logger.info("Closing with button no onscreen (close_with_button)")
    elif x == "Saving onscreen rejected" and good_list != {}:
        # Rejected entering onscreen data and no goods entered yet++
        write_or_add()
        logger.info("Saving onscreen rejected(close_with_button)")
        return "Saving onscreen rejected"
    elif x == "Good to be entered":
        # Entering onscreen confirmed:
        enter_good()
        write_or_add()
        logger.info("Saving onscreen entered(close_with_button)")
        return "onscreen entered"
    elif x == "Not valid data on screen":  # Onscreen entry fails due to invalid data
        logger.info("Onscreen entry fails due to invalid data(close_with_button)")
        return
    else:  # No onscreen data but goods are entered already
        write_or_add()
        logger.info("No onscreen entered the goods in the list(close_with_button)")


def check_on_screen():
    if entries[1, 1].get() and entries[2, 1].get():
    # checks if there are unsaved numbers on the screen
        logger.info("Unsaved data found on the screen (check_on_screen)")
        if messagebox.askyesno(message="Искате ли да "
                                       "запазим операцията на екрана?"):
            logger.info("Saving to file selected (check_on_screen)")
            if check_if_entries_are_numbers():
                logger.info("Data is valid and good is entered (check_on_screen)")
                # checks if the entered info is valid ( the numbers are floats) and creates goods
                return "Good to be entered"
            else:
                highlight_wrong()
                return "Not valid data on screen"
        else:
            return "Saving onscreen rejected"
    else:
        return "No data"


def check_if_entries_are_numbers():
    try:
        quantity = float(entries[1, 1].get().replace(",", "."))
        price = float(entries[2, 1].get().replace(",", "."))
        if quantity > 1000000.00:
            messagebox.showinfo(message="Максимално количество на операция е 1000000,00")
            logger.info("Quantity over 10000,00")
            return False
        elif price > 1000000.00:
            messagebox.showinfo(message="Максимално единична цена е 1000000,00")
            logger.info("Price over 1000000,00")
            return False
        return True
    except ValueError:
        messagebox.showinfo(title="Резултат от операция", message="Полета количество и "  
                            "цена трябва да са валидни числа")
        logger.info("not valid numbers found (check_if_entries_are_numbers)")
        highlight_wrong()
        return False


def highlight_wrong():
    # highlights the incorrect entry
    logger.info("highlights the incorrect entry(highlight_wrong)")
    entries2 = entries
    entries1 = entries2
    for q in entries1.keys():
        if entries1[q].get():
            try:
                float(entries1[q].get()) < 1000000
                entries1[q].configure(highlightthickness="0")
            except ValueError:
                entries1[q].configure(highlightthickness="2",
                                      highlightcolor="red", highlightbackground="red")
            entries1[0, 1].configure(highlightthickness="0")
    return "Wrong entries highlighted"


def enter_good():
    """Enters good not existing in inventory or adds quantity to existing"""
    if entries[1, 1].get() and entries[(2, 1)].get():
        logger.info("Quantity and price fields contain info (enter_good)")
        if check_if_entries_are_numbers() is False:
            logger.info("Entry of good failed as the numbers are not valid(enter_good)")
            return "Failed"
        else:
            # creates the good or adds quantity to existing
            create_good()
            logger.info("Good entered successfully(enter_good)")
            return "Good entered"
        for q in range(len(entries)):
            entries[q, 1].delete(0, "end")
            entries[q, 1].configure(highlightthickness="0")
        logger.info("Formatting and content cleared (enter_good)")


def create_good():
    if entries[0, 1].get():
        logger.info("Checking for name entry  (create_good)")
        # checks if name will be entered
        a = entries[0, 1].get()
        if a not in good_list.keys():
            logger.info("Checking if name exist  (create_good)")
            key = a
            a = Goods(a, float(entries[1, 1].get().replace(",", ".")),
                      float(entries[(2, 1)].get().replace(",", ".")))
            good_list[key] = a
            logger.info("Good created and name added to list(create_good)")
            m = f" created {a.name, a.quantity, a.price, a.total_price}"

        else:
            q = add_quantity()
            logger.info("Quantity added for existing good  (create_good)")
            m = q

    else:
        good = Goods("No Name good", float(entries[(1, 1)].get().replace(",", ".")),
                     float(entries[(2, 1)].get().replace(",", ".")))
        key = (x for x in range(1000000) if x not in good_list)
        good_list[key] = good
        logger.info("Good 'No Name good' created and name added to list (create_good)")
        m = f" No name created {good.name, good.quantity, good.price, good.total_price}"

    for q in range(len(entries)):
        entries[q, 1].delete(0, "end")
        logger.info(f"Entry box {entries[q, 1]} cleaned (create_good)")

    return m


def add_quantity():
    a = entries[0, 1].get()
    b = entries[1, 1].get()
    if v.get() == 1:
        # depending on choice accumulate or not the quantity for existing good.
        logger.info("Adding accumulated quantity(add_quantity)")
        if a in good_list:
                good_list[a].quantity = float(b) + good_list[a].quantity
                good_list[a].total_price = good_list[a].quantity * good_list[a].price
                logger.info(f'checked {a}')
        return f"added {good_list[a].name, good_list[a].quantity, good_list[a].price, good_list[a].total_price}"

    else:
        # creating same good but found in a different place as
        # accumulating option is not selected
        logger.info("creating same good, found in a different place"
                    " as non accumulating option is selected(add_quantity)")
        x = 1
        key = a + str(x)
        while key in good_list:
            x +=1
            key = a + str(x)
        a = Goods(a, float(entries[1, 1].get().replace(",", ".")),
                  float(entries[(2, 1)].get().replace(",", ".")))
        good_list[key] = a
        return f"added {a.name, a.quantity, a.price, a.total_price}"


def calculate_total():
    """calculate totals of all goods"""
    product_total = 0
    for i in good_list.keys():
        logger.info(f"checking {i}")
        product_total += good_list[i].total_price
    return product_total


def create_file_structure():

    wb = Workbook()
    ws = wb.active
    # ws.column_dimensions.width = 180

    ws.title = "Inventory"
    ws["A1"] = "РЕВИЗИЯ"
    ws["B1"] = "ДАТА:"
    font_1 = Font(name="Calibri", size=11, color="FFFF00")
    fill_1 = PatternFill(fill_type="solid")

    ws["A1"].font = font_1
    ws["B1"].font = font_1
    ws["A1"].fill = fill_1
    ws["B1"].fill = fill_1

    list_head_names = ["СТОКА", "БРОЙ", "ЦЕНА", "ОБЩО", "РЕЗУЛТАТ"]
    font_2 = Font(name="Calibri", size=20, underline="single")
    fill_2 = PatternFill(fill_type="solid", start_color="ADFF2F")
    # setting column and headings and patterns
    logger.info("setting column and headings and patterns (create_file_structure)")
    for r in (ws["A2":"E2"]):
        for n, c in enumerate(r):
            c.fill = fill_2
            c.font = font_2
            c.value = list_head_names[n]
    logger.info("Loop for formatting and header names (create_file_structure)")

    ws["E3"] = "=SUM(D3:D526)"
    ws["E3"].font = Font(name="Calibri", size=28)
    for i, r in enumerate(ws["A3":"E3"]):
        for n, c in enumerate(r):
            c.fill = PatternFill(fill_type="solid", start_color="B22222")

    ws["F4"] = "СТАРА РЕВИЗИЯ"
    ws["F6"] = "СТОКОВИ ОТ КОЧАН"
    ws["F8"] = "СТОКОВИ ОТ КОМПЮТЪР"
    ws["F10"] = "ОТЧЕТИ"
    ws["F12"] = "НОВА РЕВИЗИЯ"
    ws["F13"] = "=E3"
    ws["F14"] = "КРАЕН РЕЗУЛТАТ"
    ws["F14"].font = Font(name="Calibri", size=16, underline="single", italic=True, bold=True )
    ws["F15"] = "==(F13+F11)-(F5+F7+F9)"
    ws["F15"].font = Font(name="Calibri", size=48)
    ws["F15"].fill = fill_2
    logger.info("filling formulas (create_file_structure)")
    for i, r in enumerate(ws["F4":"F14"]):
        if i % 2 == 0:
            for n, c in enumerate(r):
                c.fill = fill_1
                c.font = font_1
    logger.info("Loop for column F (create_file_structure)")
    for i, r in enumerate(ws["C4":"C100"]):
        for n, c in enumerate(r):
            if c.value:
                c.fill = fill_2
    logger.info("Loop for column C (create_file_structure)")
    for col in ws.columns:
        max_length = 0
        column = utils.get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    logger.info("Loop for column width (create_file_structure)")

    for i in range(len(good_list)):
        n = str(i+4)
        ws[("A" + str(n))] = good_list[i].name
        ws[("B" + str(n))] = good_list[i].quantity
        ws["C" + str(n)] = good_list[i].price
        ws["D" + str(n)] = good_list[i].quantity*good_list[i].price
    logger.info("Loop goods attribute (create_file_structure)")
    logger.info("Saving part of the function (create_file_structure)")
    return wb


def write_to_excel():
    write_to_mysql()
    """ writes the data in a new file, ir there is existing data it is overwritten"""
    # write_to_database()
    
    if root.title() != "Програма за ревизия":
        # Destination is preselcted
        logger.info("Destination is preselcted (write_to_excel)")
        add_to_excel()
    else:
        file = select_dest_file()
        # Destination is selected
        logger.info("Destination is selected (write_to_excel)")
        if file:
            try:
                create_file_structure().save(file)
                logger.info("Check if file is open 2 (write_to_excel)")
            except :
                FileExistsError
                messagebox.showinfo(message="Файлът е отворен, затворете го за да запазите операциите ")
                logger.info("Prompted to close the file 2 (write_to_excel)")


        else:
            return "Saving canceled"
            logger.info("Canceled (write_to_excel")

    os.startfile(file)
    logger.info("Closing program (write_to_excel)")
    if __name__ == "__main__":
        sys.exit()


def select_dest_file():
    """Selects file to save the current session operations. """
    logger.info("Selecting destination file(select_dest_file)")
    file = asksaveasfilename(filetypes=[("excel", ".xlsx")], initialdir=directory,
                             initialfile=".xlsx")
    if file:
        root.title(file)
        logger.info(f"Selected destination file (select_dest_file) - {file}  ")
        return file
        # if Canceled nothing happens


def add_to_excel():
    write_to_mysql()
    """Adding to excel - keeping old records. """
    # write_to_database()
    
    file = root.title()
    wb = load_workbook(file)
    ws = wb.active
    if ws["A1"].value != "РЕВИЗИЯ":  # checks if file already have the structure
        create_file_structure()
    n = 4  # the first row to have entries after headers

    for r in ws.iter_rows(min_row=4, max_col=1):
        for c in r:  # c is cell, r is row, checking only first column as this shows if entries on the row.
            logger.info(f"Checking cell {c}, row {r} (ws.iter_rows)")
            if c.value:
                n += 1
            else:
                break  # no filled cell on the row
        else:
            logger.info("Reached continue (ws.iter_rows)")
            continue
        logger.info("Reached final break (ws.iter_rows)")
        break
    logger.info("first empty row found ")

    for i in good_list:
        ws[("A" + str(n))] = good_list[i].name
        ws[("B" + str(n))] = good_list[i].quantity
        ws["C" + str(n)] = good_list[i].price
        ws["D" + str(n)] = good_list[i].quantity*good_list[i].price
        n += 1
    logger.info(f"added rows and info in a loop starting row {n-1}")
    wb.close()
    try:
        wb.save(file)
        logger.info("Check if file is open 2 (add_to_excel)")
    except FileExistsError:
        messagebox.showinfo(message="Файлът е отворен затворете за да запазите операциите ")
        logger.info("Prompted to close the file 2 (add_to_excel)")
        return
    os.startfile(file)
    logger.info("Closing program (add_to_excel)")
    if __name__ == "__main__":
        sys.exit()


def write_to_mysql():
    ms.my_cursor.execute("create table if not exists REVISIA ( GOOD varchar(255), QUANTITY float(9,2), PRICE float(9,2), good_total float(18,2))")
    for good in good_list.keys():
         val = (str(good_list[good].name), str(good_list[good].quantity),
               str(good_list[good].price), str(good_list[good].total_price))
         ms.my_cursor.execute(ms.sql_formula, val)
         ms.mydb.commit()

    ms.my_cursor.execute("SELECT * FROM inventory.revisia")
    my_result = ms.my_cursor.fetchall()
    for i in my_result:
        print(i)
    ms.my_cursor.execute("DELETE FROM inventory.revisia")
    ms.mydb.commit()









